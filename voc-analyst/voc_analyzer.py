"""
Voice of Customer (VoC) Analyzer

Analyzes customer feedback data to extract insights, identify use cases,
and quantify demand signals for any product capability.

Works with:
- ProductBoard CSV exports
- Excel files (local or OneDrive)
- Any structured feedback data

Usage:
    from voc_analyzer import VoCAnalyzer
    
    analyzer = VoCAnalyzer()
    results = analyzer.analyze("process hierarchy", keywords=["hierarchy", "nested", "levels"])
    analyzer.generate_report(results, "process-hierarchy")
"""

import csv
import io
import json
import os
import re
import sys
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any, Optional

# Add parent paths for imports
sys.path.insert(0, str(Path(__file__).parent.parent / "outlook"))

# Try to import pandas for Excel support
try:
    import pandas as pd
    HAS_PANDAS = True
except ImportError:
    HAS_PANDAS = False

# Try to import openpyxl for Excel support without pandas
try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


class VoCAnalyzer:
    """
    Generic Voice of Customer analyzer that works for any product area.
    """
    
    # Default paths
    PRODUCTBOARD_CSV = Path(__file__).parent.parent / "productboard-insights" / "data" / "notes-export.csv"
    TEMP_DIR = Path(__file__).parent / "temp"
    OUTPUT_DIR = Path(__file__).parent.parent.parent / "outputs" / "voc-analyst"
    
    # Common text fields to search in ProductBoard exports
    SEARCHABLE_FIELDS = [
        'title', 'note', 'content', 'description', 'text', 'body',
        'customer_note', 'feedback', 'problem', 'solution', 'quote',
        'feature', 'request', 'comment', 'summary', 'details'
    ]
    
    # Fields that may contain customer/company info
    CUSTOMER_FIELDS = [
        'company', 'company_name', 'customer', 'account', 'organization',
        'org', 'client', 'user', 'owner'
    ]
    
    # Fields that may contain dates
    DATE_FIELDS = [
        'created', 'created_at', 'date', 'timestamp', 'updated', 'updated_at',
        'submitted', 'received', 'added'
    ]
    
    def __init__(self):
        """Initialize the analyzer."""
        self.data = []
        self.columns = []
        self.source_file = None
        
        # Ensure directories exist
        self.TEMP_DIR.mkdir(parents=True, exist_ok=True)
        self.OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    
    # ==================== Data Loading ====================
    
    def load_productboard_data(self, source: Optional[str] = None) -> list[dict]:
        """
        Load ProductBoard data from CSV or Excel file.
        
        Args:
            source: File path (local) or OneDrive path. If None, uses default CSV.
        
        Returns:
            List of dictionaries, one per row
        """
        if source is None:
            source = str(self.PRODUCTBOARD_CSV)
        
        source_path = Path(source)
        
        # Check if it's a local file
        if source_path.exists():
            return self._load_local_file(source_path)
        
        # Try OneDrive
        if source.startswith("onedrive:") or "/" in source:
            return self._load_onedrive_file(source.replace("onedrive:", ""))
        
        # Check default location
        if self.PRODUCTBOARD_CSV.exists():
            return self._load_local_file(self.PRODUCTBOARD_CSV)
        
        raise FileNotFoundError(f"Could not find data source: {source}")
    
    def _load_local_file(self, path: Path) -> list[dict]:
        """Load data from a local CSV or Excel file."""
        self.source_file = str(path)
        
        if path.suffix.lower() == '.csv':
            return self._load_csv(path)
        elif path.suffix.lower() in ['.xlsx', '.xls']:
            return self._load_excel(path)
        else:
            # Try CSV by default
            return self._load_csv(path)
    
    def _load_csv(self, path: Path) -> list[dict]:
        """Load data from CSV file."""
        data = []
        
        # Try different encodings
        encodings = ['utf-8', 'utf-8-sig', 'latin-1', 'cp1252']
        
        for encoding in encodings:
            try:
                with open(path, 'r', encoding=encoding, newline='') as f:
                    # Detect delimiter
                    sample = f.read(4096)
                    f.seek(0)
                    
                    sniffer = csv.Sniffer()
                    try:
                        dialect = sniffer.sniff(sample, delimiters=',;\t|')
                    except csv.Error:
                        dialect = csv.excel
                    
                    reader = csv.DictReader(f, dialect=dialect)
                    self.columns = reader.fieldnames or []
                    
                    for row in reader:
                        # Clean up row - normalize empty values
                        clean_row = {
                            k: (v.strip() if isinstance(v, str) else v)
                            for k, v in row.items()
                            if k is not None
                        }
                        data.append(clean_row)
                
                self.data = data
                print(f"✓ Loaded {len(data)} rows from {path.name}")
                print(f"  Columns: {', '.join(self.columns[:5])}{'...' if len(self.columns) > 5 else ''}")
                return data
                
            except UnicodeDecodeError:
                continue
            except Exception as e:
                if encoding == encodings[-1]:
                    raise
                continue
        
        raise ValueError(f"Could not decode CSV file: {path}")
    
    def _load_excel(self, path: Path) -> list[dict]:
        """Load data from Excel file."""
        if HAS_PANDAS:
            return self._load_excel_pandas(path)
        elif HAS_OPENPYXL:
            return self._load_excel_openpyxl(path)
        else:
            raise ImportError("pandas or openpyxl required to read Excel files. Install with: pip install pandas openpyxl")
    
    def _load_excel_pandas(self, path: Path) -> list[dict]:
        """Load Excel using pandas."""
        df = pd.read_excel(path, engine='openpyxl')
        self.columns = list(df.columns)
        
        # Convert to list of dicts
        self.data = df.fillna('').to_dict('records')
        
        print(f"✓ Loaded {len(self.data)} rows from {path.name}")
        print(f"  Columns: {', '.join(self.columns[:5])}{'...' if len(self.columns) > 5 else ''}")
        return self.data
    
    def _load_excel_openpyxl(self, path: Path) -> list[dict]:
        """Load Excel using openpyxl directly."""
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        
        # First row is headers
        self.columns = [str(h) if h else f'col_{i}' for i, h in enumerate(rows[0])]
        
        data = []
        for row in rows[1:]:
            row_dict = {}
            for i, value in enumerate(row):
                if i < len(self.columns):
                    row_dict[self.columns[i]] = str(value) if value else ''
            data.append(row_dict)
        
        wb.close()
        self.data = data
        
        print(f"✓ Loaded {len(data)} rows from {path.name}")
        print(f"  Columns: {', '.join(self.columns[:5])}{'...' if len(self.columns) > 5 else ''}")
        return data
    
    def _load_onedrive_file(self, file_path: str) -> list[dict]:
        """Load file from OneDrive using Graph API."""
        try:
            from outlook_api import get_file_content
        except ImportError:
            raise ImportError("outlook_api module required for OneDrive access")
        
        print(f"📥 Downloading from OneDrive: {file_path}")
        
        # Get file content as bytes
        content = get_file_content(file_path=file_path)
        
        # Determine file type and parse
        if file_path.lower().endswith('.csv'):
            return self._parse_csv_bytes(content)
        elif file_path.lower().endswith(('.xlsx', '.xls')):
            return self._parse_excel_bytes(content)
        else:
            # Try CSV first
            try:
                return self._parse_csv_bytes(content)
            except:
                return self._parse_excel_bytes(content)
    
    def _parse_csv_bytes(self, content: bytes) -> list[dict]:
        """Parse CSV from bytes."""
        encodings = ['utf-8', 'utf-8-sig', 'latin-1', 'cp1252']
        
        for encoding in encodings:
            try:
                text = content.decode(encoding)
                reader = csv.DictReader(io.StringIO(text))
                self.columns = reader.fieldnames or []
                self.data = [dict(row) for row in reader]
                
                print(f"✓ Loaded {len(self.data)} rows from OneDrive")
                return self.data
            except (UnicodeDecodeError, csv.Error):
                continue
        
        raise ValueError("Could not parse CSV data")
    
    def _parse_excel_bytes(self, content: bytes) -> list[dict]:
        """Parse Excel from bytes."""
        if HAS_PANDAS:
            df = pd.read_excel(io.BytesIO(content), engine='openpyxl')
            self.columns = list(df.columns)
            self.data = df.fillna('').to_dict('records')
        elif HAS_OPENPYXL:
            wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            
            if rows:
                self.columns = [str(h) if h else f'col_{i}' for i, h in enumerate(rows[0])]
                self.data = []
                for row in rows[1:]:
                    row_dict = {self.columns[i]: str(v) if v else '' for i, v in enumerate(row) if i < len(self.columns)}
                    self.data.append(row_dict)
            wb.close()
        else:
            raise ImportError("pandas or openpyxl required for Excel files")
        
        print(f"✓ Loaded {len(self.data)} rows from OneDrive")
        return self.data
    
    # ==================== Search & Analysis ====================
    
    def search_feedback(self, keywords: list[str], data: Optional[list[dict]] = None) -> list[dict]:
        """
        Search across all text fields for keyword matches.
        
        Args:
            keywords: List of keywords/phrases to search for
            data: Data to search (uses self.data if None)
        
        Returns:
            List of matching rows with match metadata
        """
        if data is None:
            data = self.data
        
        if not data:
            raise ValueError("No data loaded. Call load_productboard_data() first.")
        
        matches = []
        
        # Build regex pattern for keywords (case-insensitive)
        patterns = [re.compile(rf'\b{re.escape(kw)}\b', re.IGNORECASE) for kw in keywords]
        
        # Also create a looser pattern for partial matches
        loose_patterns = [re.compile(re.escape(kw), re.IGNORECASE) for kw in keywords]
        
        # Find searchable columns in this dataset
        searchable_cols = self._find_columns(self.SEARCHABLE_FIELDS)
        
        for row in data:
            # Combine all searchable text
            text_parts = []
            for col in searchable_cols:
                value = row.get(col, '')
                if value:
                    text_parts.append(str(value))
            
            combined_text = ' '.join(text_parts)
            
            # Check for matches
            matched_keywords = []
            match_count = 0
            
            for i, pattern in enumerate(patterns):
                if pattern.search(combined_text):
                    matched_keywords.append(keywords[i])
                    match_count += len(pattern.findall(combined_text))
            
            # Also check loose matches if no exact matches
            if not matched_keywords:
                for i, pattern in enumerate(loose_patterns):
                    if pattern.search(combined_text):
                        matched_keywords.append(keywords[i] + " (partial)")
                        match_count += 1
            
            if matched_keywords:
                match_entry = {
                    **row,
                    '_matched_keywords': matched_keywords,
                    '_match_count': match_count,
                    '_combined_text': combined_text[:500]  # First 500 chars for context
                }
                matches.append(match_entry)
        
        # Sort by match count (most relevant first)
        matches.sort(key=lambda x: x['_match_count'], reverse=True)
        
        print(f"✓ Found {len(matches)} matches for keywords: {', '.join(keywords)}")
        return matches
    
    def _find_columns(self, field_names: list[str]) -> list[str]:
        """Find columns that match given field name patterns."""
        found = []
        for col in self.columns:
            col_lower = col.lower().replace(' ', '_').replace('-', '_')
            for field in field_names:
                if field in col_lower or col_lower in field:
                    found.append(col)
                    break
        return found if found else self.columns  # Return all columns if no matches
    
    def extract_use_cases(self, matches: list[dict]) -> list[dict]:
        """
        Extract structured use case information from matches.
        
        Args:
            matches: List of matching rows from search_feedback()
        
        Returns:
            List of structured use case dictionaries
        """
        use_cases = []
        
        # Find customer and date columns
        customer_cols = self._find_columns(self.CUSTOMER_FIELDS)
        date_cols = self._find_columns(self.DATE_FIELDS)
        text_cols = self._find_columns(self.SEARCHABLE_FIELDS)
        
        for match in matches:
            # Extract customer info
            customer = None
            for col in customer_cols:
                value = match.get(col, '')
                if value and value.strip():
                    customer = str(value).strip()
                    break
            
            # Extract date
            date = None
            for col in date_cols:
                value = match.get(col, '')
                if value:
                    date = self._parse_date(value)
                    break
            
            # Extract main text/quote
            quote = None
            context = None
            for col in text_cols:
                value = match.get(col, '')
                if value and len(str(value)) > 20:
                    if quote is None:
                        quote = str(value).strip()
                    elif context is None:
                        context = str(value).strip()
            
            use_case = {
                'customer': customer or 'Unknown',
                'quote': quote or match.get('_combined_text', '')[:300],
                'context': context,
                'date': date,
                'matched_keywords': match.get('_matched_keywords', []),
                'match_count': match.get('_match_count', 1),
                'raw_data': {k: v for k, v in match.items() if not k.startswith('_')}
            }
            
            use_cases.append(use_case)
        
        print(f"✓ Extracted {len(use_cases)} use cases")
        return use_cases
    
    def _parse_date(self, value: Any) -> Optional[str]:
        """Try to parse a date value into ISO format."""
        if not value:
            return None
        
        value_str = str(value).strip()
        
        # Common date formats
        formats = [
            '%Y-%m-%d',
            '%Y-%m-%dT%H:%M:%S',
            '%Y-%m-%dT%H:%M:%SZ',
            '%Y-%m-%d %H:%M:%S',
            '%d/%m/%Y',
            '%m/%d/%Y',
            '%d-%m-%Y',
            '%B %d, %Y',
            '%b %d, %Y',
        ]
        
        for fmt in formats:
            try:
                dt = datetime.strptime(value_str[:19], fmt)
                return dt.strftime('%Y-%m-%d')
            except ValueError:
                continue
        
        # Return as-is if we can't parse
        return value_str[:10] if len(value_str) >= 10 else None
    
    def categorize_by_theme(self, use_cases: list[dict], num_categories: int = 5) -> dict[str, list[dict]]:
        """
        Group use cases by similar themes.
        
        This performs simple keyword-based clustering. For AI-based categorization,
        the LLM should analyze the use cases directly.
        
        Args:
            use_cases: List of use case dictionaries
            num_categories: Target number of categories
        
        Returns:
            Dictionary of category_name -> list of use cases
        """
        categories = defaultdict(list)
        
        # Extract common themes from matched keywords
        keyword_counts = defaultdict(int)
        for uc in use_cases:
            for kw in uc.get('matched_keywords', []):
                # Clean up keyword
                clean_kw = kw.replace(' (partial)', '').lower()
                keyword_counts[clean_kw] += 1
        
        # Use top keywords as category names
        top_keywords = sorted(keyword_counts.items(), key=lambda x: x[1], reverse=True)[:num_categories]
        
        if not top_keywords:
            # Fallback: put all in one category
            categories['General Feedback'] = use_cases
            return dict(categories)
        
        # Assign each use case to a category based on its keywords
        for uc in use_cases:
            matched = uc.get('matched_keywords', [])
            
            # Find best matching category
            best_category = 'Other'
            for kw, count in top_keywords:
                for m in matched:
                    if kw in m.lower():
                        best_category = kw.title()
                        break
                if best_category != 'Other':
                    break
            
            categories[best_category].append(uc)
        
        # Sort categories by size
        sorted_categories = dict(sorted(categories.items(), key=lambda x: len(x[1]), reverse=True))
        
        print(f"✓ Categorized into {len(sorted_categories)} themes")
        return sorted_categories
    
    def calculate_demand_signal(self, use_cases: list[dict]) -> dict:
        """
        Calculate demand signal metrics from use cases.
        
        Args:
            use_cases: List of use case dictionaries
        
        Returns:
            Dictionary with demand metrics
        """
        if not use_cases:
            return {
                'unique_customers': 0,
                'total_mentions': 0,
                'first_request': None,
                'most_recent': None,
                'customers': [],
                'trend': 'unknown'
            }
        
        # Count unique customers
        customers = set()
        for uc in use_cases:
            customer = uc.get('customer', 'Unknown')
            if customer and customer != 'Unknown':
                customers.add(customer)
        
        # Total mentions (sum of match counts)
        total_mentions = sum(uc.get('match_count', 1) for uc in use_cases)
        
        # Date range
        dates = []
        for uc in use_cases:
            date = uc.get('date')
            if date:
                dates.append(date)
        
        dates = sorted(dates) if dates else []
        first_request = dates[0] if dates else None
        most_recent = dates[-1] if dates else None
        
        # Simple trend calculation
        trend = 'stable'
        if len(dates) >= 4:
            mid = len(dates) // 2
            first_half = len(dates[:mid])
            second_half = len(dates[mid:])
            if second_half > first_half * 1.5:
                trend = 'increasing'
            elif first_half > second_half * 1.5:
                trend = 'decreasing'
        
        demand = {
            'unique_customers': len(customers),
            'total_mentions': total_mentions,
            'first_request': first_request,
            'most_recent': most_recent,
            'customers': sorted(list(customers)),
            'trend': trend
        }
        
        print(f"✓ Demand signal: {len(customers)} customers, {total_mentions} mentions")
        return demand
    
    # ==================== Main Analysis ====================
    
    def analyze(self, capability_name: str, keywords: list[str], source: Optional[str] = None) -> dict:
        """
        Run complete analysis for a capability.
        
        Args:
            capability_name: Name of the capability/feature (e.g., "process hierarchy")
            keywords: List of search keywords
            source: Data source (file path or OneDrive path)
        
        Returns:
            Complete analysis results dictionary
        """
        print(f"\n{'='*60}")
        print(f"🔍 VoC Analysis: {capability_name}")
        print(f"{'='*60}\n")
        
        # Load data
        self.load_productboard_data(source)
        
        # Search for matches
        matches = self.search_feedback(keywords)
        
        # Extract use cases
        use_cases = self.extract_use_cases(matches)
        
        # Categorize
        categories = self.categorize_by_theme(use_cases)
        
        # Calculate demand
        demand = self.calculate_demand_signal(use_cases)
        
        results = {
            'capability_name': capability_name,
            'keywords': keywords,
            'source': self.source_file,
            'analysis_date': datetime.now().strftime('%Y-%m-%d'),
            'total_records': len(self.data),
            'matches_found': len(matches),
            'use_cases': use_cases,
            'categories': categories,
            'demand_signal': demand
        }
        
        print(f"\n✓ Analysis complete for: {capability_name}")
        return results
    
    # ==================== Report Generation ====================
    
    def generate_report(self, results: dict, output_name: Optional[str] = None) -> str:
        """
        Generate a markdown report from analysis results.
        
        Args:
            results: Analysis results from analyze()
            output_name: Output filename (without extension)
        
        Returns:
            Path to generated report
        """
        capability = results['capability_name']
        date = results['analysis_date']
        
        if output_name is None:
            output_name = capability.lower().replace(' ', '-')
        
        filename = f"{output_name}-{date}.md"
        output_path = self.OUTPUT_DIR / filename
        
        # Build report
        report = []
        
        # Header
        report.append(f"# VoC Analysis: {capability.title()}")
        report.append(f"\n*Generated: {date}*\n")
        
        # Executive Summary
        demand = results['demand_signal']
        report.append("## Executive Summary\n")
        report.append(f"- **{demand['unique_customers']}** unique customers mentioned this capability")
        report.append(f"- **{demand['total_mentions']}** total mentions across {results['total_records']} records")
        report.append(f"- **{len(results['categories'])}** distinct use case themes identified")
        report.append(f"- **Keywords searched**: {', '.join(results['keywords'])}")
        if results['source']:
            report.append(f"- **Data source**: {Path(results['source']).name}")
        report.append("")
        
        # Demand Signal
        report.append("## Demand Signal\n")
        report.append(f"| Metric | Value |")
        report.append(f"|--------|-------|")
        report.append(f"| Unique Customers | {demand['unique_customers']} |")
        report.append(f"| Total Mentions | {demand['total_mentions']} |")
        report.append(f"| First Request | {demand['first_request'] or 'Unknown'} |")
        report.append(f"| Most Recent | {demand['most_recent'] or 'Unknown'} |")
        report.append(f"| Trend | {demand['trend'].title()} |")
        report.append("")
        
        # Customer List
        if demand['customers']:
            report.append("### Customers Requesting This Capability\n")
            for customer in demand['customers'][:20]:  # Limit to 20
                report.append(f"- {customer}")
            if len(demand['customers']) > 20:
                report.append(f"- *...and {len(demand['customers']) - 20} more*")
            report.append("")
        
        # Use Case Categories
        report.append("## Use Case Categories\n")
        
        categories = results['categories']
        for category, use_cases in categories.items():
            report.append(f"### {category}")
            report.append(f"**Count: {len(use_cases)} mentions**\n")
            
            # Create table
            report.append("| Customer | Quote | Date |")
            report.append("|----------|-------|------|")
            
            for uc in use_cases[:10]:  # Limit to 10 per category
                customer = uc.get('customer', 'Unknown')[:30]
                quote = uc.get('quote', '')[:100].replace('|', '/').replace('\n', ' ')
                date = uc.get('date', 'N/A')
                report.append(f"| {customer} | {quote}... | {date} |")
            
            if len(use_cases) > 10:
                report.append(f"\n*...and {len(use_cases) - 10} more in this category*")
            report.append("")
        
        # Recommendations
        report.append("## Recommendations\n")
        report.append("Based on this analysis:\n")
        
        if demand['unique_customers'] >= 10:
            report.append("- ✅ **Strong demand signal** - Multiple customers requesting this capability")
        elif demand['unique_customers'] >= 5:
            report.append("- ⚠️ **Moderate demand** - Consider for roadmap planning")
        else:
            report.append("- ℹ️ **Emerging need** - Monitor for additional requests")
        
        if demand['trend'] == 'increasing':
            report.append("- 📈 **Growing interest** - Requests are increasing over time")
        elif demand['trend'] == 'decreasing':
            report.append("- 📉 **Declining interest** - Fewer recent requests")
        
        report.append("\n---\n")
        report.append(f"*Report generated by VoC Analyst skill on {date}*")
        
        # Write report
        report_text = '\n'.join(report)
        output_path.write_text(report_text, encoding='utf-8')
        
        print(f"\n📄 Report saved: {output_path}")
        return str(output_path)
    
    def save_raw_data(self, results: dict, output_name: Optional[str] = None) -> str:
        """Save raw analysis data as JSON for further processing."""
        capability = results['capability_name']
        date = results['analysis_date']
        
        if output_name is None:
            output_name = capability.lower().replace(' ', '-')
        
        filename = f"{output_name}-{date}-raw.json"
        output_path = self.TEMP_DIR / filename
        
        # Prepare serializable data
        export_data = {
            'capability_name': results['capability_name'],
            'keywords': results['keywords'],
            'analysis_date': results['analysis_date'],
            'demand_signal': results['demand_signal'],
            'categories': {k: len(v) for k, v in results['categories'].items()},
            'use_cases': [
                {
                    'customer': uc.get('customer'),
                    'quote': uc.get('quote', '')[:500],
                    'date': uc.get('date'),
                    'matched_keywords': uc.get('matched_keywords', [])
                }
                for uc in results['use_cases']
            ]
        }
        
        output_path.write_text(json.dumps(export_data, indent=2), encoding='utf-8')
        print(f"💾 Raw data saved: {output_path}")
        return str(output_path)


# ==================== Convenience Functions ====================

def analyze_capability(capability_name: str, keywords: list[str], source: Optional[str] = None, generate_report: bool = True) -> dict:
    """
    Convenience function to run a complete analysis.
    
    Args:
        capability_name: Name of the capability (e.g., "process hierarchy")
        keywords: Search keywords
        source: Data source path (optional)
        generate_report: Whether to generate markdown report
    
    Returns:
        Analysis results dictionary
    """
    analyzer = VoCAnalyzer()
    results = analyzer.analyze(capability_name, keywords, source)
    
    if generate_report:
        analyzer.generate_report(results)
        analyzer.save_raw_data(results)
    
    return results


def quick_search(keywords: list[str], source: Optional[str] = None) -> list[dict]:
    """
    Quick search without full analysis - just returns matches.
    
    Args:
        keywords: Search keywords
        source: Data source path (optional)
    
    Returns:
        List of matching rows
    """
    analyzer = VoCAnalyzer()
    analyzer.load_productboard_data(source)
    return analyzer.search_feedback(keywords)


# ==================== CLI ====================

if __name__ == "__main__":
    import argparse
    
    parser = argparse.ArgumentParser(description="Voice of Customer Analyzer")
    parser.add_argument("capability", help="Capability name to analyze")
    parser.add_argument("--keywords", "-k", nargs="+", required=True, help="Search keywords")
    parser.add_argument("--source", "-s", help="Data source (file path or OneDrive path)")
    parser.add_argument("--no-report", action="store_true", help="Skip report generation")
    
    args = parser.parse_args()
    
    results = analyze_capability(
        capability_name=args.capability,
        keywords=args.keywords,
        source=args.source,
        generate_report=not args.no_report
    )
    
    print(f"\n{'='*60}")
    print(f"Summary: {results['demand_signal']['unique_customers']} customers, "
          f"{results['demand_signal']['total_mentions']} mentions")
    print(f"{'='*60}")