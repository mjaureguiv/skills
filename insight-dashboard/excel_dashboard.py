"""
========================================
EXCEL DASHBOARD GENERATOR
========================================
This script processes customer insights from a CSV/Excel file and generates
a comprehensive Excel dashboard with:
- Summary metrics
- Cluster distribution charts
- Feature ranking with RICE scores
- Detailed breakdown by cluster
- Sample customer quotes

Usage:
    python excel_dashboard.py <input_file> [output_file]
    
Example:
    python excel_dashboard.py insights.xlsx dashboard_output.xlsx
"""

import pandas as pd
import sys
import os
from datetime import datetime
from pathlib import Path

# Import our existing modules
from config import CLUSTER_DEFINITIONS, EFFORT_BY_CLUSTER, DEFAULT_EFFORT
from data_processor import load_data, prepare_insights
from classifier import classify_all_insights
from rice_scorer import score_all_features, score_subfeatures

# Try to import xlsxwriter, fall back to openpyxl
try:
    import xlsxwriter
    USE_XLSXWRITER = True
except ImportError:
    USE_XLSXWRITER = False
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import PieChart, BarChart, Reference
    from openpyxl.chart.label import DataLabelList
    from openpyxl.utils.dataframe import dataframe_to_rows


def create_excel_dashboard_xlsxwriter(classified_df, features_df, subfeatures_df, output_path):
    """Create Excel dashboard using xlsxwriter (better charts)."""
    
    workbook = xlsxwriter.Workbook(output_path)
    
    # Define formats
    title_format = workbook.add_format({
        'bold': True, 'font_size': 24, 'font_color': '#0070F2',
        'align': 'center', 'valign': 'vcenter'
    })
    header_format = workbook.add_format({
        'bold': True, 'font_size': 12, 'bg_color': '#0070F2', 
        'font_color': 'white', 'align': 'center', 'valign': 'vcenter',
        'border': 1
    })
    metric_label_format = workbook.add_format({
        'bold': True, 'font_size': 11, 'align': 'center', 
        'valign': 'vcenter', 'bg_color': '#E8E8E8'
    })
    metric_value_format = workbook.add_format({
        'bold': True, 'font_size': 18, 'font_color': '#0070F2',
        'align': 'center', 'valign': 'vcenter'
    })
    section_format = workbook.add_format({
        'bold': True, 'font_size': 14, 'font_color': '#333333',
        'bottom': 2, 'bottom_color': '#0070F2'
    })
    cell_format = workbook.add_format({
        'align': 'left', 'valign': 'vcenter', 'border': 1
    })
    number_format = workbook.add_format({
        'align': 'center', 'valign': 'vcenter', 'border': 1, 'num_format': '#,##0'
    })
    percent_format = workbook.add_format({
        'align': 'center', 'valign': 'vcenter', 'border': 1, 'num_format': '0.0%'
    })
    rice_high_format = workbook.add_format({
        'align': 'center', 'valign': 'vcenter', 'border': 1,
        'bg_color': '#90EE90', 'bold': True
    })
    rice_medium_format = workbook.add_format({
        'align': 'center', 'valign': 'vcenter', 'border': 1,
        'bg_color': '#FFFFE0', 'bold': True
    })
    rice_low_format = workbook.add_format({
        'align': 'center', 'valign': 'vcenter', 'border': 1,
        'bg_color': '#FFB6C1', 'bold': True
    })
    quote_format = workbook.add_format({
        'italic': True, 'text_wrap': True, 'valign': 'top',
        'font_color': '#555555', 'border': 1
    })
    
    # =========================================================================
    # SHEET 1: DASHBOARD OVERVIEW
    # =========================================================================
    dashboard = workbook.add_worksheet('Dashboard')
    dashboard.set_column('A:A', 5)
    dashboard.set_column('B:G', 18)
    dashboard.set_column('H:H', 5)
    
    # Title
    dashboard.merge_range('B2:G2', '🎯 Customer Insight Intelligence Dashboard', title_format)
    dashboard.write('B3', f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M")}')
    
    # Key Metrics
    row = 5
    dashboard.write(f'B{row}', 'KEY METRICS', section_format)
    row += 2
    
    total_insights = len(classified_df)
    unique_companies = classified_df['cleaned_company'].nunique()
    num_clusters = len(features_df)
    avg_rice = features_df['rice_score'].mean()
    
    metrics = [
        ('Total Insights', f'{total_insights:,}'),
        ('Companies Impacted', f'{unique_companies:,}'),
        ('Feature Clusters', str(num_clusters)),
        ('Avg RICE Score', f'{avg_rice:.1f}')
    ]
    
    col = 1  # Column B
    for label, value in metrics:
        dashboard.write(row, col, label, metric_label_format)
        dashboard.write(row + 1, col, value, metric_value_format)
        col += 1
    
    row += 4
    
    # Cluster Distribution Chart Data
    dashboard.write(f'B{row}', 'DISTRIBUTION BY CLUSTER', section_format)
    row += 2
    
    cluster_counts = classified_df['cluster'].value_counts()
    chart_data_start_row = row
    
    dashboard.write(row - 1, 1, 'Cluster', header_format)
    dashboard.write(row - 1, 2, 'Insights', header_format)
    dashboard.write(row - 1, 3, 'Companies', header_format)
    dashboard.write(row - 1, 4, '% of Total', header_format)
    
    for cluster in cluster_counts.index:
        count = cluster_counts[cluster]
        companies = classified_df[classified_df['cluster'] == cluster]['cleaned_company'].nunique()
        pct = count / total_insights
        
        dashboard.write(row, 1, cluster, cell_format)
        dashboard.write(row, 2, count, number_format)
        dashboard.write(row, 3, companies, number_format)
        dashboard.write(row, 4, pct, percent_format)
        row += 1
    
    chart_data_end_row = row - 1
    
    # Create pie chart
    pie_chart = workbook.add_chart({'type': 'pie'})
    pie_chart.add_series({
        'name': 'Distribution by Cluster',
        'categories': f'=Dashboard!$B${chart_data_start_row}:$B${chart_data_end_row}',
        'values': f'=Dashboard!$C${chart_data_start_row}:$C${chart_data_end_row}',
        'data_labels': {'percentage': True, 'category': False},
    })
    pie_chart.set_title({'name': 'Insights by Cluster'})
    pie_chart.set_style(10)
    pie_chart.set_size({'width': 400, 'height': 300})
    dashboard.insert_chart(f'F{chart_data_start_row - 1}', pie_chart)
    
    # Create bar chart for companies
    bar_chart = workbook.add_chart({'type': 'bar'})
    bar_chart.add_series({
        'name': 'Companies per Cluster',
        'categories': f'=Dashboard!$B${chart_data_start_row}:$B${chart_data_end_row}',
        'values': f'=Dashboard!$D${chart_data_start_row}:$D${chart_data_end_row}',
        'fill': {'color': '#0070F2'},
    })
    bar_chart.set_title({'name': 'Companies by Cluster'})
    bar_chart.set_style(11)
    bar_chart.set_size({'width': 400, 'height': 300})
    bar_chart.set_legend({'none': True})
    dashboard.insert_chart(f'F{chart_data_start_row + 12}', bar_chart)
    
    # =========================================================================
    # SHEET 2: FEATURE RANKING
    # =========================================================================
    ranking = workbook.add_worksheet('Feature Ranking')
    ranking.set_column('A:A', 6)
    ranking.set_column('B:B', 30)
    ranking.set_column('C:C', 40)
    ranking.set_column('D:G', 14)
    ranking.set_column('H:H', 12)
    
    ranking.merge_range('A1:H1', '🏆 Feature Ranking by RICE Score', title_format)
    ranking.set_row(0, 40)
    
    # Headers
    headers = ['Rank', 'Feature', 'Sub-features', 'Insights', 'Companies', '% of Total', 'RICE Score']
    for col, header in enumerate(headers):
        ranking.write(2, col, header, header_format)
    
    # Data rows
    row = 3
    for _, feat in features_df.iterrows():
        ranking.write(row, 0, int(feat['rank']), number_format)
        ranking.write(row, 1, feat['feature_name'], cell_format)
        ranking.write(row, 2, feat['subfeatures'], cell_format)
        ranking.write(row, 3, int(feat['insight_count']), number_format)
        ranking.write(row, 4, int(feat['company_count']), number_format)
        ranking.write(row, 5, feat['company_percentage'] / 100, percent_format)
        
        # Color code RICE score
        rice = feat['rice_score']
        if rice >= 10:
            rice_format = rice_high_format
        elif rice >= 5:
            rice_format = rice_medium_format
        else:
            rice_format = rice_low_format
        ranking.write(row, 6, rice, rice_format)
        row += 1
    
    # RICE Score chart
    rice_chart = workbook.add_chart({'type': 'bar'})
    rice_chart.add_series({
        'name': 'RICE Score',
        'categories': f'=\'Feature Ranking\'!$B$4:$B${row}',
        'values': f'=\'Feature Ranking\'!$G$4:$G${row}',
        'fill': {'color': '#1A9898'},
        'data_labels': {'value': True},
    })
    rice_chart.set_title({'name': 'Feature Priority (RICE Scores)'})
    rice_chart.set_style(11)
    rice_chart.set_size({'width': 500, 'height': 350})
    rice_chart.set_legend({'none': True})
    ranking.insert_chart('I3', rice_chart)
    
    # =========================================================================
    # SHEET 3: FEATURE DETAILS
    # =========================================================================
    details = workbook.add_worksheet('Feature Details')
    details.set_column('A:A', 25)
    details.set_column('B:B', 20)
    details.set_column('C:C', 12)
    details.set_column('D:D', 12)
    details.set_column('E:E', 60)
    details.set_column('F:F', 25)
    
    details.merge_range('A1:F1', '📋 Feature Details with Sample Quotes', title_format)
    details.set_row(0, 40)
    
    headers = ['Cluster', 'Sub-cluster', 'Insights', 'Companies', 'Sample Quote', 'Company']
    for col, header in enumerate(headers):
        details.write(2, col, header, header_format)
    
    row = 3
    for cluster in features_df['feature_name']:
        cluster_df = classified_df[classified_df['cluster'] == cluster]
        
        # Get top quotes per subcluster
        for subcluster in cluster_df['subcluster'].unique():
            sub_df = cluster_df[cluster_df['subcluster'] == subcluster]
            insight_count = len(sub_df)
            company_count = sub_df['cleaned_company'].nunique()
            
            # Get a sample quote (longest one)
            if 'text_length' in sub_df.columns:
                sample_row = sub_df.nlargest(1, 'text_length').iloc[0]
            else:
                sample_row = sub_df.iloc[0]
            
            quote = sample_row.get('cleaned_text', sample_row.get('note_text', ''))
            company = sample_row.get('cleaned_company', sample_row.get('company_name', ''))
            
            # Truncate quote if too long
            if len(str(quote)) > 300:
                quote = str(quote)[:297] + '...'
            
            details.write(row, 0, cluster, cell_format)
            details.write(row, 1, subcluster, cell_format)
            details.write(row, 2, insight_count, number_format)
            details.write(row, 3, company_count, number_format)
            details.write(row, 4, quote, quote_format)
            details.write(row, 5, company, cell_format)
            details.set_row(row, 45)  # Taller row for quotes
            row += 1
    
    # =========================================================================
    # SHEET 4: RICE BREAKDOWN
    # =========================================================================
    rice_sheet = workbook.add_worksheet('RICE Breakdown')
    rice_sheet.set_column('A:A', 30)
    rice_sheet.set_column('B:G', 14)
    
    rice_sheet.merge_range('A1:G1', '📊 RICE Score Breakdown', title_format)
    rice_sheet.set_row(0, 40)
    
    headers = ['Feature', 'Reach', 'Impact', 'Confidence', 'Effort', 'RICE Score']
    for col, header in enumerate(headers):
        rice_sheet.write(2, col, header, header_format)
    
    row = 3
    for _, feat in features_df.iterrows():
        rice_sheet.write(row, 0, feat['feature_name'], cell_format)
        rice_sheet.write(row, 1, int(feat['reach']), number_format)
        rice_sheet.write(row, 2, feat['impact'], number_format)
        rice_sheet.write(row, 3, feat['confidence'], number_format)
        rice_sheet.write(row, 4, feat['effort'], number_format)
        
        rice = feat['rice_score']
        if rice >= 10:
            rice_format = rice_high_format
        elif rice >= 5:
            rice_format = rice_medium_format
        else:
            rice_format = rice_low_format
        rice_sheet.write(row, 5, rice, rice_format)
        row += 1
    
    # Add formula explanation
    row += 2
    rice_sheet.write(row, 0, 'RICE Formula:', section_format)
    row += 1
    rice_sheet.write(row, 0, 'RICE Score = (Reach × Impact × Confidence) / Effort')
    row += 2
    rice_sheet.write(row, 0, 'Legend:', section_format)
    row += 1
    rice_sheet.write(row, 0, 'Reach = Number of unique companies')
    row += 1
    rice_sheet.write(row, 0, 'Impact = Severity (0.5-3.0) based on language')
    row += 1
    rice_sheet.write(row, 0, 'Confidence = Data confidence (0-1.0)')
    row += 1
    rice_sheet.write(row, 0, 'Effort = Estimated complexity (1-10)')
    
    # =========================================================================
    # SHEET 5: ALL CLASSIFIED DATA
    # =========================================================================
    raw_data = workbook.add_worksheet('Classified Data')
    
    # Select relevant columns
    export_cols = ['insight_id', 'cluster', 'subcluster', 'cleaned_text', 
                   'cleaned_company', 'cluster_confidence', 'overall_confidence']
    export_cols = [c for c in export_cols if c in classified_df.columns]
    
    # Also include original columns if available
    for orig_col in ['note_text', 'company_name', 'created_at']:
        if orig_col in classified_df.columns and orig_col not in export_cols:
            export_cols.append(orig_col)
    
    export_df = classified_df[export_cols]
    
    # Headers
    for col, header in enumerate(export_df.columns):
        raw_data.write(0, col, header, header_format)
    
    # Data
    for row_idx, (_, data_row) in enumerate(export_df.iterrows(), start=1):
        for col_idx, value in enumerate(data_row):
            if pd.isna(value):
                raw_data.write(row_idx, col_idx, '')
            else:
                raw_data.write(row_idx, col_idx, str(value)[:500])  # Truncate long text
    
    # Auto-filter
    raw_data.autofilter(0, 0, len(export_df), len(export_cols) - 1)
    
    # Set column widths
    raw_data.set_column('A:A', 10)
    raw_data.set_column('B:C', 25)
    raw_data.set_column('D:D', 60)
    raw_data.set_column('E:E', 25)
    raw_data.set_column('F:G', 12)
    
    workbook.close()
    print(f"\n✅ Excel dashboard created: {output_path}")


def create_excel_dashboard_openpyxl(classified_df, features_df, subfeatures_df, output_path):
    """Create Excel dashboard using openpyxl (fallback)."""
    
    wb = Workbook()
    
    # Styles
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='0070F2', end_color='0070F2', fill_type='solid')
    title_font = Font(bold=True, size=20, color='0070F2')
    
    # =========================================================================
    # SHEET 1: DASHBOARD
    # =========================================================================
    ws = wb.active
    ws.title = 'Dashboard'
    
    ws.merge_cells('B2:G2')
    ws['B2'] = '🎯 Customer Insight Intelligence Dashboard'
    ws['B2'].font = title_font
    ws['B2'].alignment = Alignment(horizontal='center')
    
    ws['B3'] = f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M")}'
    
    # Key Metrics
    ws['B5'] = 'KEY METRICS'
    ws['B5'].font = Font(bold=True, size=14)
    
    metrics = [
        ('Total Insights', len(classified_df)),
        ('Companies Impacted', classified_df['cleaned_company'].nunique()),
        ('Feature Clusters', len(features_df)),
        ('Avg RICE Score', round(features_df['rice_score'].mean(), 1))
    ]
    
    for col, (label, value) in enumerate(metrics, start=2):
        ws.cell(row=7, column=col, value=label).font = Font(bold=True)
        ws.cell(row=8, column=col, value=value).font = Font(bold=True, size=16, color='0070F2')
    
    # Cluster Distribution
    ws['B10'] = 'DISTRIBUTION BY CLUSTER'
    ws['B10'].font = Font(bold=True, size=14)
    
    cluster_counts = classified_df['cluster'].value_counts()
    headers = ['Cluster', 'Insights', 'Companies', '% of Total']
    for col, header in enumerate(headers, start=2):
        cell = ws.cell(row=12, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
    
    row = 13
    total = len(classified_df)
    for cluster, count in cluster_counts.items():
        companies = classified_df[classified_df['cluster'] == cluster]['cleaned_company'].nunique()
        ws.cell(row=row, column=2, value=cluster)
        ws.cell(row=row, column=3, value=count)
        ws.cell(row=row, column=4, value=companies)
        ws.cell(row=row, column=5, value=f'{count/total*100:.1f}%')
        row += 1
    
    # =========================================================================
    # SHEET 2: FEATURE RANKING
    # =========================================================================
    ws2 = wb.create_sheet('Feature Ranking')
    
    ws2.merge_cells('A1:G1')
    ws2['A1'] = '🏆 Feature Ranking by RICE Score'
    ws2['A1'].font = title_font
    
    headers = ['Rank', 'Feature', 'Sub-features', 'Insights', 'Companies', '% of Total', 'RICE Score']
    for col, header in enumerate(headers, start=1):
        cell = ws2.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
    
    row = 4
    for _, feat in features_df.iterrows():
        ws2.cell(row=row, column=1, value=int(feat['rank']))
        ws2.cell(row=row, column=2, value=feat['feature_name'])
        ws2.cell(row=row, column=3, value=feat['subfeatures'])
        ws2.cell(row=row, column=4, value=int(feat['insight_count']))
        ws2.cell(row=row, column=5, value=int(feat['company_count']))
        ws2.cell(row=row, column=6, value=f"{feat['company_percentage']:.1f}%")
        
        rice_cell = ws2.cell(row=row, column=7, value=feat['rice_score'])
        if feat['rice_score'] >= 10:
            rice_cell.fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
        elif feat['rice_score'] >= 5:
            rice_cell.fill = PatternFill(start_color='FFFFE0', end_color='FFFFE0', fill_type='solid')
        else:
            rice_cell.fill = PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid')
        
        row += 1
    
    # =========================================================================
    # SHEET 3: CLASSIFIED DATA
    # =========================================================================
    ws3 = wb.create_sheet('Classified Data')
    
    export_cols = ['insight_id', 'cluster', 'subcluster', 'cleaned_text', 
                   'cleaned_company', 'overall_confidence']
    export_cols = [c for c in export_cols if c in classified_df.columns]
    
    for col, header in enumerate(export_cols, start=1):
        cell = ws3.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
    
    for row_idx, (_, data_row) in enumerate(classified_df[export_cols].iterrows(), start=2):
        for col_idx, col_name in enumerate(export_cols, start=1):
            value = data_row[col_name]
            if pd.notna(value):
                ws3.cell(row=row_idx, column=col_idx, value=str(value)[:500])
    
    wb.save(output_path)
    print(f"\n✅ Excel dashboard created: {output_path}")


def generate_dashboard(input_file: str, output_file: str = None, tag_filter: str = None):
    """
    Main function to generate the Excel dashboard.
    
    Parameters:
        input_file: Path to input CSV/Excel file with insights
        output_file: Optional output path. If not provided, creates dashboard in same folder.
        tag_filter: Optional tag to filter insights (e.g., "Process Governance")
    """
    print("=" * 60)
    print("🎯 Customer Insight Intelligence - Excel Dashboard Generator")
    print("=" * 60)
    
    # Determine output path
    if output_file is None:
        input_path = Path(input_file)
        output_file = str(input_path.parent / f"insight_dashboard_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
    
    # Step 1: Load data
    print("\n📁 Loading data...")
    raw_df = load_data(input_file)
    
    # Step 1b: Filter by tag if specified
    if tag_filter:
        print(f"\n🏷️ Filtering for tag: '{tag_filter}'...")
        if 'tags' in raw_df.columns:
            original_count = len(raw_df)
            raw_df = raw_df[raw_df['tags'].str.contains(tag_filter, na=False, case=False)]
            print(f"   Filtered from {original_count:,} to {len(raw_df):,} insights")
        else:
            print("   ⚠️ No 'tags' column found, skipping filter")
    
    # Step 2: Prepare data
    print("\n🔄 Preparing insights...")
    prepared_df = prepare_insights(raw_df)
    
    # Step 3: Classify insights
    print("\n🤖 Classifying insights into clusters...")
    classified_df = classify_all_insights(prepared_df)
    
    # Step 4: Calculate RICE scores
    print("\n📊 Calculating RICE scores...")
    features_df = score_all_features(classified_df)
    subfeatures_df = score_subfeatures(classified_df)
    
    # Step 5: Create Excel dashboard
    print("\n📈 Creating Excel dashboard...")
    if USE_XLSXWRITER:
        create_excel_dashboard_xlsxwriter(classified_df, features_df, subfeatures_df, output_file)
    else:
        create_excel_dashboard_openpyxl(classified_df, features_df, subfeatures_df, output_file)
    
    print(f"\n✅ Dashboard generation complete!")
    print(f"📄 Output file: {output_file}")
    print("\nSheets included:")
    print("  1. Dashboard - Summary metrics and charts")
    print("  2. Feature Ranking - Prioritized features with RICE scores")
    print("  3. Feature Details - Breakdown with sample quotes")
    print("  4. RICE Breakdown - Detailed scoring components")
    print("  5. Classified Data - All insights with classifications")
    
    return output_file


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python excel_dashboard.py <input_file> [output_file] [--tag-filter=TAG]")
        print("\nExample:")
        print("  python excel_dashboard.py insights.xlsx")
        print("  python excel_dashboard.py insights.csv my_dashboard.xlsx")
        print("  python excel_dashboard.py insights.csv output.xlsx --tag-filter=\"Process Governance\"")
        sys.exit(1)
    
    input_file = sys.argv[1]
    output_file = None
    tag_filter = None
    
    # Parse arguments
    for arg in sys.argv[2:]:
        if arg.startswith("--tag-filter="):
            tag_filter = arg.split("=", 1)[1].strip('"').strip("'")
        elif not arg.startswith("--"):
            output_file = arg
    
    generate_dashboard(input_file, output_file, tag_filter)
